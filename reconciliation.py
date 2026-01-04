import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from columns_to_keep import COLUMNS_GL
from columns_to_keep import COLUMNS_TB
from gl_processing import DR_LEFT, CR_LEFT

START_ROW = 4  # leaves rows 1–3 free for headers
SHEET_NAME = "TB&GL Reconciliation"

def build_recon_skeleton(tb_df):
    return pd.DataFrame({
        "Account": extract_needed_accounts(tb_df),
        "Description": None,
        "Movement DR (TB)": None,
        "Movement CR (TB)": None,
        "Movement DR (GL)": None,
        "Movement CR (GL)": None,
        "Check DR": None,
        "Check CR": None,
    })

def extract_needed_accounts(tb_df):
  acc_str = tb_df[COLUMNS_TB["acc"]].str.strip()


  numeric_acc = pd.to_numeric(acc_str, errors='coerce')

  # accounts that are exactly 4 digits and not ending with 00
  mask1 = (acc_str.str.len() == 4) & \
          (acc_str.str.isdigit()) & \
          (numeric_acc % 100 != 0)

  # accounts that start with letters: eg: B600
  mask2 = (acc_str.str.len() == 4) & (acc_str.str[:1].str.isalpha())
  mask = mask1 | mask2

  return acc_str[mask]

def format_excel(ws, recon_df: pd.DataFrame, company_name) -> None:    # Add header
    ws["A1"] = f"TB & GL Reconciliation of {company_name}" if company_name else "TB & GL Reconciliation"
    ws["A1"].font = Font(name='Sylfaen', size=12, bold=True)
    ws["A2"] = "Reporting Date:"

    #Set column widths
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

# column helper
def col_letter(df, col_name):
    return get_column_letter(df.columns.get_loc(col_name) + 1)

def get_tb_movemenet(ws, direction: str, tb_df, account_col: str, tb_m_col: str, r: int) -> None:
    # TB MovementColumn formula
    m_col = col_letter(tb_df, COLUMNS_TB[f"movement_{direction}"])
    tb_m_formula = f'=SUMIFS(TB!${m_col}:${m_col},TB!$A:$A,{account_col}{r})'
    ws[f"{tb_m_col}{r}"] = tb_m_formula

def get_gl_movement(ws, direction: str, gl_df, account_col: str, target_col: str, r: int) -> None:
    # Pick the correct column in GL based on direction
    # DO not use GL_COLUMNS bc those left columns are manually added
    if direction.lower() == "dr":
        gl_account_col = col_letter(gl_df, DR_LEFT)
    elif direction.lower() == "cr":
        gl_account_col = col_letter(gl_df, CR_LEFT)

    # GL Amount column
    gl_amount_col = col_letter(gl_df, COLUMNS_GL["amount"])

    gl_formula = (
        f'=SUMIFS(GL!${gl_amount_col}:${gl_amount_col},'
        f'GL!${gl_account_col}:${gl_account_col},{account_col}{r})'
    )

    ws[f"{target_col}{r}"] = gl_formula

def add_reconciliation_formulas(ws, recon_df: pd.DataFrame, tb_df, gl_df) -> None:
	#Pre define styles
	sylfaen_font = Font(name='Sylfaen', size=10)
	num_format = '#,##0;[Red](#,##0);-'

	# Apply styles to data rows
	first_data_row = START_ROW + 1
	last_data_row = START_ROW + len(recon_df)
	print("FIRST DATA ROW:", first_data_row, "\n last data row:", last_data_row)

	account_col = col_letter(recon_df, "Account")
	desc_col = col_letter(recon_df, "Description")

	tb_dr_col = col_letter(recon_df, "Movement DR (TB)")
	tb_cr_col = col_letter(recon_df, "Movement CR (TB)")
	gl_dr_col = col_letter(recon_df, "Movement DR (GL)")
	gl_cr_col = col_letter(recon_df, "Movement CR (GL)")

	check_dr_col = col_letter(recon_df, "Check DR")
	check_cr_col = col_letter(recon_df, "Check CR")
	
	numeric_cols = [tb_dr_col, tb_cr_col, gl_dr_col, gl_cr_col, check_dr_col, check_cr_col]
	
	for r in range(first_data_row, last_data_row + 1):
			# Description formula: VLOOKUP from TB first
			tb_acc_col  = col_letter(tb_df, COLUMNS_TB["acc"])
			tb_desc_col = col_letter(tb_df, COLUMNS_TB["name"])
			tb_desc_idx = tb_df.columns.get_loc(COLUMNS_TB["name"]) - tb_df.columns.get_loc(COLUMNS_TB["acc"]) + 1
			desc_formula = f'=IFERROR(VLOOKUP({account_col}{r},TB!${tb_acc_col}:${tb_desc_col},{tb_desc_idx}, FALSE),0)'
			ws[f"{desc_col}{r}"] = desc_formula

			# TB Movement formulas
			get_tb_movemenet(ws, "dr", tb_df, account_col, tb_dr_col, r)
			get_tb_movemenet(ws, "cr", tb_df, account_col, tb_cr_col, r)

			get_gl_movement(ws, "dr", gl_df, account_col,  gl_dr_col, r)
			get_gl_movement(ws, "cr", gl_df, account_col,  gl_cr_col, r)

			# Check movements
			ws[f"{check_dr_col}{r}"] = f"={tb_dr_col}{r}-{gl_dr_col}{r}"
			ws[f"{check_cr_col}{r}"] = f"={tb_cr_col}{r}-{gl_cr_col}{r}"
      
			# Apply Sylfaen 12 to the whole row
			for col_idx in range(1, len(recon_df.columns) + 1):
					cell = ws.cell(row=r, column=col_idx)
					cell.font = sylfaen_font
			
			# Apply Number Format only to the numeric columns
			for col_let in numeric_cols:
					ws[f"{col_let}{r}"].number_format = num_format


def reconcile_data(tb_df: pd.DataFrame, gl_df: pd.DataFrame, writer: pd.ExcelWriter, company_name) -> pd.DataFrame:

    # 1. Build the reconciliation base table (structure only)
    recon_df = build_recon_skeleton(tb_df)

    # 2. Write raw data
    recon_df.to_excel(
      writer,
      sheet_name=SHEET_NAME,
      index=False,
      startrow=START_ROW - 1  # pandas is 0-based
    )

    ws = writer.sheets[SHEET_NAME]
    format_excel(ws, recon_df, company_name)


    # 3. Inject Excel formulas
    add_reconciliation_formulas(ws, recon_df, tb_df, gl_df)
    return recon_df # test
